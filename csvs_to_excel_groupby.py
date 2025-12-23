#!/usr/bin/env python3

from pathlib import Path
import argparse
import sys
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# =============================
# CONFIG
# =============================

INPUT_KEEP_COLS = [
    "DNS", "Title", "Type", "Port", "Protocol", "CVE ID",
    "Vendor Reference", "CVSS3.1 Base", "Threat", "Impact",
    "Solution", "Results"
]

FINAL_COLS_TEMPLATE = [
    "GROUP_FIELD", "Title", "Port", "Protocol", "CVE ID", "Vendor Reference",
    "CVSS3.1 Base", "Severity", "Threat", "Impact",
    "Solution", "Results"
]

DNS_ALIASES = ["DNS", "FQDN", "Hostname", "Host", "Host Name"]
IP_ALIASES = ["IP", "IP Address", "IPAddress", "Host IP", "IPv4", "IPv6"]

CVSS_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)")

DEFAULT_COL_WIDTHS = {
    "DNS": 32,
    "IP": 18,
    "Title": 60,
    "Port": 8,
    "Protocol": 10,
    "CVE ID": 14,
    "Vendor Reference": 28,
    "CVSS3.1 Base": 14,
    "Severity": 12,
    "Threat": 20,
    "Impact": 20,
    "Solution": 40,
    "Results": 30,
    "Rows": 8,
    "Critical": 9,
    "High": 9,
    "Medium": 9,
    "Low": 9,
    "__default__": 18
}

# =============================
# HELPERS
# =============================

def parse_cvss_score(text):
    if pd.isna(text):
        return None
    m = CVSS_RE.search(str(text))
    return float(m.group(1)) if m else None

def cvss_to_severity(score):
    if score is None:
        return "Unknown"
    if score == 0.0:
        return "None"
    if score <= 3.9:
        return "Low"
    if score <= 6.9:
        return "Medium"
    if score <= 8.9:
        return "High"
    return "Critical"

def sanitize_sheet_name(name):
    s = str(name)
    for ch in [':', '\\', '/', '?', '*', '[', ']']:
        s = s.replace(ch, '_')
    return s.strip()[:31] or "Host"

def detect_group_column(df, mode):
    aliases = DNS_ALIASES if mode == "dns" else IP_ALIASES
    lower = {c.lower(): c for c in df.columns}
    for a in aliases:
        if a.lower() in lower:
            return lower[a.lower()]
    return None

# =============================
# CORE LOGIC
# =============================

def collect_all_hosts(folder, mode):
    hosts = set()
    for f in folder.glob("*.csv"):
        try:
            df = pd.read_csv(f, skiprows=7, dtype=str)
        except Exception:
            continue
        col = detect_group_column(df, mode)
        if not col:
            hosts.add("<UNKNOWN>")
            continue
        vals = df[col].dropna().astype(str).str.strip()
        for v in vals:
            hosts.add(v or "<UNKNOWN>")
    return sorted(hosts)

def process_csvs(folder, mode):
    host_data = {}

    for f in folder.glob("*.csv"):
        df = pd.read_csv(f, skiprows=7, dtype=str)

        for c in INPUT_KEEP_COLS:
            if c not in df.columns:
                df[c] = pd.NA

        group_col = detect_group_column(df, mode)
        if not group_col:
            df["_GROUP_"] = "<UNKNOWN>"
            group_col = "_GROUP_"

        df["Type"] = df["Type"].astype(str).str.upper()
        df = df[df["Type"] == "VULN"]
        if df.empty:
            continue

        df["Severity"] = df["CVSS3.1 Base"].apply(
            lambda x: cvss_to_severity(parse_cvss_score(x))
        )

        df = df.drop(columns=["Type"])
        df = df.rename(columns={group_col: "GROUP_FIELD"})
        df["GROUP_FIELD"] = df["GROUP_FIELD"].fillna("<UNKNOWN>")

        for host, g in df.groupby("GROUP_FIELD"):
            host_data.setdefault(host, []).append(g)

    return {k: pd.concat(v, ignore_index=True) for k, v in host_data.items()}

def write_excel(all_hosts, data, output, widths, mode):
    host_col = "DNS" if mode == "dns" else "IP"
    final_cols = [c if c != "GROUP_FIELD" else host_col for c in FINAL_COLS_TEMPLATE]

    host_to_sheet = {}

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        rows = []
        for h in all_hosts:
            df = data.get(h)
            if df is None:
                rows.append({host_col: h, "Rows": 0, "Critical": 0, "High": 0, "Medium": 0, "Low": 0})
            else:
                vc = df["Severity"].value_counts()
                rows.append({
                    host_col: h,
                    "Rows": len(df),
                    "Critical": vc.get("Critical", 0),
                    "High": vc.get("High", 0),
                    "Medium": vc.get("Medium", 0),
                    "Low": vc.get("Low", 0)
                })

        covered = pd.DataFrame(rows)
        covered.to_excel(writer, sheet_name="Covered Hosts", index=False)

        for host, df in data.items():
            sheet = sanitize_sheet_name(host)
            i = 1
            while sheet in writer.sheets:
                sheet = f"{sheet[:28]}_{i}"
                i += 1

            host_to_sheet[host] = sheet

            df = df.rename(columns={"GROUP_FIELD": host_col})
            for c in final_cols:
                if c not in df.columns:
                    df[c] = pd.NA
            df[final_cols].to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(output)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for r in ws.iter_rows():
            for c in r:
                c.border = border

        for i, cell in enumerate(ws[1], start=1):
            w = widths.get(cell.value, widths.get("__default__", 18))
            ws.column_dimensions[get_column_letter(i)].width = w

    # hyperlinks
    ws = wb["Covered Hosts"]
    col_idx = [c.value for c in ws[1]].index(host_col) + 1
    for r in range(2, ws.max_row + 1):
        host = ws.cell(r, col_idx).value
        if host in host_to_sheet:
            ws.cell(r, col_idx).hyperlink = f"#{host_to_sheet[host]}!A1"
            ws.cell(r, col_idx).style = "Hyperlink"

    wb.save(output)

# =============================
# ENTRY POINT
# =============================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-i", "--input", help="CSV folder")
    ap.add_argument("-o", "--output", help="Output Excel file")
    ap.add_argument("-g", "--group-by", choices=["dns", "ip"])
    ap.add_argument("-w", "--widths", help="JSON column width map")
    args = ap.parse_args()

    folder = Path(args.input or input("CSV folder: ").strip())
    output = Path(args.output or input("Output Excel file: ").strip())

    mode = args.group_by or input("Group by DNS or IP? [dns]: ").strip().lower() or "dns"

    widths = DEFAULT_COL_WIDTHS.copy()
    if args.widths:
        widths.update(json.loads(Path(args.widths).read_text() if Path(args.widths).exists() else args.widths))

    all_hosts = collect_all_hosts(folder, mode)
    data = process_csvs(folder, mode)
    write_excel(all_hosts, data, output, widths, mode)

    print(f"[+] Excel report created: {output}")

if __name__ == "__main__":
    main()
